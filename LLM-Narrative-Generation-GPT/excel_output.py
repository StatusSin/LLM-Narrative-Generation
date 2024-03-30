import openpyxl
from openpyxl.styles import Alignment

def append_to_excel(file_name, data, row_height=400, column_width=20):
    # Load the workbook and select the first sheet
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active

    # Append data
    row = sheet.max_row + 1  # Get the next row number
    sheet.append(data)

    # Adjust row height and column widths
    sheet.row_dimensions[row].height = 409  # Maximum height
    for col in range(1, 14):  # Assuming you're using columns A to M
        col_letter = openpyxl.utils.get_column_letter(col)
        sheet.column_dimensions[col_letter].width = column_width

        # Apply text wrap to each cell in this row
        cell = sheet[f'{col_letter}{row}']
        cell.alignment = Alignment(wrap_text=True)

    # Merge cells from B to M for the response
    sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=13)

    # Save the workbook
    workbook.save(file_name)